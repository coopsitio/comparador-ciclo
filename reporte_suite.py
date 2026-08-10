"""
reporte_suite.py - Reporte CONSOLIDADO de una revision (suite) de un ciclo.

Lee los pares v7_*.csv / v8_*.csv de local/suite_<pefa>/, compara cada dimension y
arma UN Excel (resumen_<pefa>.xlsx) con:
  - Hoja RESUMEN: una fila por dimension (diferencias, solo-V7, solo-V8, mayor
    brecha) + deducciones automaticas.
  - Una hoja por dimension: tabla con V7, V8, Delta (V8-V7), %Delta y tipo,
    ordenada por mayor diferencia absoluta.

Uso:  python reporte_suite.py --pefa 15281
"""
import argparse
import csv
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AZUL = PatternFill("solid", fgColor="1F4E78")
GRIS = PatternFill("solid", fgColor="D9E1F2")
ROJO = PatternFill("solid", fgColor="F8CBAD")
BLANCO_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
MAX_FILAS = 3000  # tope de filas por hoja de dimension (las dimensiones por producto pueden ser grandes)


def cargar(ruta):
    with open(ruta, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def a_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def clave(row):
    return (row.get("empresa", ""), row.get("cuenta", ""), row.get("concepto", ""))


def comparar(v7_rows, v8_rows):
    """Devuelve lista de dicts con las diferencias, ordenada por |delta| desc."""
    v7 = {clave(r): a_num(r.get("monto")) for r in v7_rows}
    v8 = {clave(r): a_num(r.get("monto")) for r in v8_rows}
    difs = []
    for k in set(v7) | set(v8):
        a, b = v7.get(k), v8.get(k)
        if a == b:
            continue
        if a is None:
            tipo, delta, pct = "solo en V8", b, None
        elif b is None:
            tipo, delta, pct = "solo en V7", -a, None
        else:
            tipo = "distinto"
            delta = b - a
            pct = (delta / a * 100) if a else None
        difs.append({"empresa": k[0], "cuenta": k[1], "concepto": k[2],
                     "v7": a, "v8": b, "delta": delta, "pct": pct, "tipo": tipo})
    difs.sort(key=lambda d: abs(d["delta"]) if d["delta"] is not None else 0, reverse=True)
    return difs


def nombre_dim(base):
    # "01_facturacion_por_tarifa_concepto" -> "facturacion por tarifa concepto"
    partes = base.split("_")
    if partes and partes[0].isdigit():
        partes = partes[1:]
    return " ".join(partes)


def deduccion(difs):
    """Frase corta con el hallazgo principal de la dimension."""
    if not difs:
        return "Sin diferencias: V7 y V8 cuadran."
    solo7 = sum(1 for d in difs if d["tipo"] == "solo en V7")
    solo8 = sum(1 for d in difs if d["tipo"] == "solo en V8")
    top = difs[0]
    partes = []
    if solo7:
        partes.append(f"{solo7} solo en V7")
    if solo8:
        partes.append(f"{solo8} solo en V8")
    detalle = f"mayor brecha: {top['concepto']}"
    if top["cuenta"]:
        detalle = f"mayor brecha: {top['cuenta']} / {top['concepto']}"
    if top["v7"] is not None and top["v8"] is not None:
        detalle += f" (V7 {top['v7']:,.0f} vs V8 {top['v8']:,.0f})"
    partes.append(detalle)
    return "; ".join(partes)


def estilar_encabezado(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = AZUL
        cell.font = BLANCO_BOLD
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pefa", required=True)
    ap.add_argument("--carpeta", default=None)
    args = ap.parse_args()

    raiz = os.path.dirname(os.path.abspath(__file__))
    carpeta = args.carpeta or os.path.join(raiz, "local", f"suite_{args.pefa}")
    v7_files = sorted(glob.glob(os.path.join(carpeta, "v7_*.csv")))

    wb = Workbook()
    resumen = wb.active
    resumen.title = "RESUMEN"
    resumen.append(["Dimension", "Filas V7", "Filas V8", "Diferencias",
                    "Solo V7", "Solo V8", "Distinto", "Hallazgo principal"])

    for v7f in v7_files:
        base = os.path.basename(v7f)[3:-4]  # quita "v7_" y ".csv"
        v8f = os.path.join(carpeta, "v8_" + base + ".csv")
        if not os.path.exists(v8f):
            continue
        v7_rows, v8_rows = cargar(v7f), cargar(v8f)
        difs = comparar(v7_rows, v8_rows)
        solo7 = sum(1 for d in difs if d["tipo"] == "solo en V7")
        solo8 = sum(1 for d in difs if d["tipo"] == "solo en V8")
        dist = sum(1 for d in difs if d["tipo"] == "distinto")

        hallazgo = deduccion(difs)
        if len(difs) > MAX_FILAS:
            hallazgo = f"[hoja muestra top {MAX_FILAS} de {len(difs)}] " + hallazgo
        resumen.append([nombre_dim(base), len(v7_rows), len(v8_rows), len(difs),
                        solo7, solo8, dist, hallazgo])

        # Hoja de la dimension (nombre de hoja <= 31 chars); tope de filas por si es grande
        hoja = wb.create_sheet(title=base[:31])
        hoja.append(["empresa", "cuenta", "concepto", "V7", "V8",
                     "Delta (V8-V7)", "%Delta", "tipo"])
        for d in difs[:MAX_FILAS]:
            hoja.append([
                d["empresa"], d["cuenta"], d["concepto"],
                d["v7"], d["v8"], d["delta"],
                round(d["pct"], 1) if d["pct"] is not None else None,
                d["tipo"],
            ])
        estilar_encabezado(hoja, 8)
        for col, w in zip("ABCDEFGH", (12, 22, 26, 14, 14, 14, 10, 12)):
            hoja.column_dimensions[col].width = w
        # resalta filas solo-en
        for r in range(2, hoja.max_row + 1):
            if hoja.cell(row=r, column=8).value != "distinto":
                hoja.cell(row=r, column=8).fill = ROJO

    estilar_encabezado(resumen, 8)
    for col, w in zip("ABCDEFGH", (34, 10, 10, 12, 9, 9, 10, 70)):
        resumen.column_dimensions[col].width = w
    for r in range(2, resumen.max_row + 1):
        resumen.cell(row=r, column=1).font = BOLD

    salida = os.path.join("local", f"suite_{args.pefa}", f"resumen_{args.pefa}.xlsx")
    wb.save(os.path.join(raiz, salida))
    print(f"Reporte consolidado: {salida}")


if __name__ == "__main__":
    main()
